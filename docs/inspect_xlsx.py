"""원본 학습데이터 설계서.xlsx 구조를 정확하게 출력 (한글 보존)"""
import openpyxl
import sys

p = "c:/lsc/Kftc/학습데이터 설계서.xlsx"
wb = openpyxl.load_workbook(p, data_only=False)

sys.stdout.reconfigure(encoding="utf-8")

print("=" * 70)
print("Sheets:", wb.sheetnames)
print("=" * 70)

for s in wb.sheetnames:
    ws = wb[s]
    print(f"\n┌──── [{s}] ────")
    print(f"│ dims={ws.dimensions} max_row={ws.max_row} max_col={ws.max_column}")
    print(f"└─── merged: {[str(r) for r in ws.merged_cells.ranges][:10]}")
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=min(20, ws.max_row), values_only=True), 1):
        cells = []
        for v in row:
            if v is None:
                cells.append("")
            else:
                s2 = str(v).replace("\n", "↩")[:50]
                cells.append(s2)
        if any(c.strip() for c in cells):
            print(f"  [r{r:>2}] " + " | ".join(cells))
